"""訂單匯出 Excel 服務"""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def export_orders_to_excel(group, orders) -> BytesIO:
    """匯出團單訂單為 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "訂單明細"
    
    # 樣式
    header_font = Font(bold=True, color="5B4733")
    header_fill = PatternFill(start_color="C9A977", end_color="C9A977", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 標題資訊
    ws.merge_cells('A1:F1')
    ws['A1'] = f"團單：{group.name}"
    ws['A1'].font = Font(bold=True, size=14)
    
    ws.merge_cells('A2:F2')
    store_name = group.store.name if group.store else (group.store_name or "（店家已移除）")
    ws['A2'] = f"店家：{store_name} | 截止：{group.deadline.strftime('%Y/%m/%d %H:%M')}"
    ws['A2'].font = Font(size=10, color="666666")
    
    # 表頭
    headers = ["訂購人", "品項", "數量", "規格", "備註", "小計"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 訂單資料
    row = 5
    total_amount = 0
    
    for order in orders:
        if order.status.value != 'submitted':
            continue
            
        for item in order.items:
            ws.cell(row=row, column=1, value=order.user.show_name).border = thin_border
            ws.cell(row=row, column=2, value=item.item_name).border = thin_border
            ws.cell(row=row, column=3, value=item.quantity).border = thin_border
            
            # 規格（size/sugar/ice 是欄位，選項與加料是關聯物件）
            specs = []
            if item.size:
                specs.append(item.size)
            if item.sugar:
                specs.append(item.sugar)
            if item.ice:
                specs.append(item.ice)
            for opt in item.selected_options:
                specs.append(opt.option_name)
            for top in item.selected_toppings:
                specs.append(f"+{top.topping_name}")
            ws.cell(row=row, column=4, value=" / ".join(specs) if specs else "-").border = thin_border
            
            ws.cell(row=row, column=5, value=item.note or "-").border = thin_border
            
            # 小計用 OrderItem.subtotal（已含單價+選項+加料+數量），與個人/店家明細一致
            subtotal = float(item.subtotal)
            ws.cell(row=row, column=6, value=subtotal).border = thin_border
            ws.cell(row=row, column=6).number_format = '$#,##0'
            
            total_amount += subtotal
            row += 1

        # 折扣列（有折扣才加）
        if order.discount_amount and order.discount_amount > 0:
            ws.cell(row=row, column=1, value=order.user.show_name).border = thin_border
            note = f"折扣（{order.discount_note}）" if order.discount_note else "折扣"
            ws.cell(row=row, column=2, value=note).border = thin_border
            ws.cell(row=row, column=4, value="-").border = thin_border
            ws.cell(row=row, column=5, value="-").border = thin_border
            disc = -float(order.discount_amount)
            ws.cell(row=row, column=6, value=disc).border = thin_border
            ws.cell(row=row, column=6).number_format = '$#,##0'
            ws.cell(row=row, column=6).font = Font(color="C0392B")
            total_amount += disc
            row += 1
    
    # 合計
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row=row, column=1, value="合計").font = Font(bold=True)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=6, value=total_amount).font = Font(bold=True)
    ws.cell(row=row, column=6).number_format = '$#,##0'
    
    # 外送費
    if group.delivery_fee:
        row += 1
        ws.merge_cells(f'A{row}:E{row}')
        ws.cell(row=row, column=1, value="外送費").alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=6, value=float(group.delivery_fee)).number_format = '$#,##0'
        
        row += 1
        ws.merge_cells(f'A{row}:E{row}')
        ws.cell(row=row, column=1, value="總計").font = Font(bold=True)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=6, value=total_amount + float(group.delivery_fee)).font = Font(bold=True)
        ws.cell(row=row, column=6).number_format = '$#,##0'
    
    # 調整欄寬
    column_widths = [12, 25, 8, 20, 15, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # 儲存到 BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output
